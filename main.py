import os
import re
from tkinter import NO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

from domain import check_app_version
from enum import Enum

#todo: забрать версию проекта и на ее основе создать имя файла эксель
#todo: 

class Param:
    def __init__(self):
        self.code = 0  # Класс числовой
        self.name = ""  # название текстовое
        self.value = None  # значение пустое
        self.units = ""  # Единицы измерения текстовое
        self.type = ""  # тип строковый может быть "UNION", "UINT16", "INT16", "STR", "TIME", "DATE"
        self.view = ""   # Вид строковый может быть "DEC", "HEX", "BIN"
        self.address = 0  # адрес числовой
        self.record = 0  # Запись числовой
        self.min_value = 0  # Минимум числовой
        self.max_value = 0  # максимум числовой
        self.factory_setting = 0  # Заводская уставка числовой
        self.coefficient = 0  # коэффициент числовой
        self.size = 0  # размер числовой всегла 2
        self.rows = ""  # Строки текстовый
        self.description = ""  # Описание текстовый
        self.hidden = 0  # скрытый числовой
        self.method = ""  # Методика проверки текстовый
        self.comments_user = ""  # Комментарии для пользователя текстовый
        self.comments_dev = ""  # Комментарии для разработчика текстовый
        self.recommendations = ""  # Рекомендации (что делать, если не работает) текстовый
        self.additional = ""  # Дополнительно текстовый
        self.date_changed = ""  # Дата изменения текстовый
        self.author = ""  # Автор текстовый

def create_excel_file(excel_file, param_objects, groups):
    """
    Создание файла Excel на основе массива объектов Param и данных о группах.

    Args:
        excel_file (str): Путь к выходному файлу Excel.
        param_objects (list): Список объектов Param.
        groups (list): Список групп.
    """
    # Создаем новую книгу Excel
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Parameters"

    # Заголовки столбцов
    headers = [
        "Код",
        "Название",
        "Значение",
        "Ед. изм.",
        "Тип",
        "Вид",
        "Адрес",
        "Запись",
        "Мин.",
        "Макс.",
        "Завод. установ.",
        "Коэф-т",
        "Размер",
        "Строки",
        "Описание стр. значений/битов",
        "Скрытый",
        "Описание",
        "Комментарии для пользователя",
        "Методика проверки",
        "Комментарии для разработчика",
        "Рекомендации (что делать, если не работает)",
        "Дополнительно",
        "Дата изменения",
        "Автор"
    ]

    # Записываем заголовки в первую строку
    for i, header in enumerate(headers):
        cell = sheet.cell(row=1, column=i + 1)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="67a0ef", end_color="67a0ef", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Заполняем данные
    row_index = 2  # Начинаем с второй строки (первая строка — заголовки)

    # Группируем параметры по группам
    grouped_params = {}
    for param in param_objects:
        group_code = param.code.split(".")[0]  # Извлекаем номер группы из кода
        if group_code not in grouped_params:
            grouped_params[group_code] = []
        grouped_params[group_code].append(param)

    # Заливка для строки с группой (салатовый цвет)
    group_fill = PatternFill(start_color="09c850", end_color="09c850", fill_type="solid")

    # Записываем группы и их параметры
    for group in groups:
        group_number = group["group_number"]
        group_index = group["group_index"]
        group_description = group["group_description"]

        # Записываем номер группы и её название в столбцы 1 и 2
        sheet.cell(row=row_index, column=1, value=group_number)
        sheet.cell(row=row_index, column=2, value= group_index + ". " + group_description)

        # Подкрашиваем строку с группой салатовым цветом
        for col in range(1, len(headers) + 1):
            sheet.cell(row=row_index, column=col).fill = group_fill

        row_index += 1

        # Записываем параметры этой группы
        if group_number in grouped_params:
            for param in grouped_params[group_number]:
                sheet.cell(row=row_index, column=1, value=param.code)
                sheet.cell(row=row_index, column=2, value=param.name)
                sheet.cell(row=row_index, column=3, value=param.value)
                sheet.cell(row=row_index, column=4, value=param.units)
                sheet.cell(row=row_index, column=5, value=param.type)
                sheet.cell(row=row_index, column=6, value=param.view)
                sheet.cell(row=row_index, column=7, value=param.address)
                sheet.cell(row=row_index, column=8, value=param.record)
                sheet.cell(row=row_index, column=9, value=param.min_value)
                sheet.cell(row=row_index, column=10, value=param.max_value)
                sheet.cell(row=row_index, column=11, value=param.factory_setting)
                sheet.cell(row=row_index, column=12, value=param.coefficient)
                sheet.cell(row=row_index, column=13, value=param.size)
                sheet.cell(row=row_index, column=14, value=param.rows)
                sheet.cell(row=row_index, column=15, value=param.description)
                sheet.cell(row=row_index, column=16, value=param.hidden)
                sheet.cell(row=row_index, column=17, value=param.method)
                sheet.cell(row=row_index, column=18, value=param.comments_user)
                sheet.cell(row=row_index, column=19, value=param.comments_dev)
                sheet.cell(row=row_index, column=20, value=param.recommendations)
                sheet.cell(row=row_index, column=21, value=param.additional)
                sheet.cell(row=row_index, column=22, value=param.date_changed)
                sheet.cell(row=row_index, column=23, value=param.author)
                row_index += 1

    # Автоматическое выравнивание ширины столбцов
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Получаем букву столбца
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width

    # Сохраняем файл
    wb.save(excel_file)
    print(f"Файл {excel_file} успешно создан.")
            
# ready
def find_version_file():
    """
    Поиск файла 'project_version.h' или 'config.h' в текущем каталоге и его
    подкаталогах.

    Returns:
        str: Путь найденного файла или None, если файл не найден
    """
    filenames = ["project_version.h", "config.h"]
    for filename in filenames:
        # Walk through the directory tree
        for root, dirs, files in os.walk("."):
            # Check if the current filename is in the list of files
            if filename in files:
                return os.path.join(root, filename)

    # File not found in any directory
    print("Error: Neither 'project_version.h' nor 'config.h' found")
    return None
    
def find_param_file():
    """
    Поиск файла 'params.h' или 'menu_params.h' в текущем каталоге и его
    подкаталогах.

    Returns:
        str: Путь найденного файла или None, если файл не найден
    """
    filenames = ["params.h", "menu_params.h"]
    for filename in filenames:
        # Walk through the directory tree
        for root, dirs, files in os.walk("."):
            # Check if the current filename is in the list of files
            if filename in files:
                return os.path.join(root, filename)

    # File not found in any directory
    print("Error: Neither 'params.h' nor 'menu_params.h' found")
    return None
# ready
def parse_strings(file_path, start_line_number):
    with open(file_path, 'r', encoding='Windows-1251') as file:
        content = file.read()

    # Находим блок между //! Строки и //! Строки конец
    start_marker = "//! Строки"
    end_marker = "//! Строки конец"
    start_index = content.find(start_marker)
    end_index = content.find(end_marker)

    if start_index == -1 or end_index == -1:
        print("Error: String markers not found")
        return []

    string_block = content[start_index:end_index]

    # Удаляем блоки между #if BUR_M и #else
    string_block = re.sub(r'#if BUR_M.*?#else', '', string_block, flags=re.DOTALL)
    # Удаляем оставшиеся #if BUR_M, #else, #endif
    string_block = re.sub(r'#if BUR_M|#else|#endif', '', string_block)

    # Разделяем на участки по //!----------------
    sections = re.split(r'//!----------------', string_block)

    result = []
    found_start = False  # Флаг, чтобы начать добавление строк после нахождения start_line_number

    for section in sections:

        # Ищем все строки в формате "   НЕТ АВАРИЙ   ", // 35
        matches = re.findall(r'"(.*?)",\s*//\s*(\d+)', section)
        print(matches)
        current_index = 0  # Нумерация с 0 для каждого раздела

        for match in matches:
            text, line_number = match
            line_number = int(line_number)  # Преобразуем номер строки в число

            # Если нашли строку с номером start_line_number или больше, начинаем добавлять
            if line_number >= start_line_number:
                found_start = True

            if found_start:
                result.append(f"{current_index}-{text.strip()}")  # Добавляем точку с запятой
                current_index += 1

        # Если нашли строки, добавляем разделитель и завершаем обработку
        if found_start:
            break

    return result

def parse_groups(file_path):
    """
    Парсит группы между //! Группы и //! Группы конец.
    Преобразует строки в формат "ГРУППА X НАЗВАНИЕ".

    Args:
        file_path (str): Путь к файлу.

    Returns:
        list: Список групп в формате "ГРУППА X НАЗВАНИЕ".
    """
    with open(file_path, 'r', encoding='Windows-1251') as file:
        content = file.read()

    # Находим блок между //! Группы и //! Группы конец
    start_marker = "//! Группы"
    end_marker = "//! Группы конец"
    start_index = content.find(start_marker)
    end_index = content.find(end_marker)

    if start_index == -1 or end_index == -1:
        print("Error: Group markers not found")
        return []

    group_block = content[start_index:end_index]

    # Ищем все строки в формате "   4 ГРУППА D   ", "    КОМАНДЫ     ", GR_INIT(GroupD, 0),
    matches = re.findall(r'\"\s*(\d+)\s*ГРУППА\s*(\w+)\s*\"\s*,\s*\"\s*([^\"]+)\s*\"', group_block)

    groups = []
    for match in matches:
        group_number, group_name, group_description = match
        # Формируем строку в формате "ГРУППА X НАЗВАНИЕ"
        group_dict = {
            "group_number": group_number,
            "group_index": group_name,
            "group_description": group_description.strip()
        }
        groups.append(group_dict)

    return groups

def parse_parameters(file_path):
    """
    Парсит параметры между //! Параметры и //! Параметры конец.
    Группирует параметры по группам и извлекает:
    - имя
    - размерность
    - минимальное значение
    - максимальное значение
    - значение по умолчанию
    - кодировку (обрабатывается отдельно)
    - адрес параметра

    Args:
        file_path (str): Путь к файлу.

    Returns:
        dict: Словарь, где ключ — название группы, а значение — список параметров.
              Каждый параметр представлен словарем с полями:
              - group_index: индекс группы (например, "Т")
              - param_number: номер параметра (например, "0")
              - param_name: имя параметра (например, "ТЕХНОЛ. РЕГ.")
              - unit: размерность
              - min_value: минимальное значение
              - max_value: максимальное значение
              - default_value: значение по умолчанию
              - encoding: кодировка (сырая строка, для дальнейшей обработки)
              - address: адрес параметра
    """
    with open(file_path, 'r', encoding='Windows-1251') as file:
        content = file.read()

    # Находим блок между //! Параметры и //! Параметры конец
    start_marker = "//! Параметры"
    end_marker = "//! Параметры конец"
    start_index = content.find(start_marker)
    end_index = content.find(end_marker)

    if start_index == -1 or end_index == -1:
        print("Error: Parameter markers not found")
        return {}

    parameter_block = content[start_index:end_index]

    # Регулярное выражение для поиска групп и параметров
    group_pattern = re.compile(r'//! ГРУППА (\w+) ([^\n]+)')
    param_pattern = re.compile(
        r'\"([^\"]+)\"\s*,\s*\"([^\"]*)\"\s*,\s*([^,]+)\s*,\s*([^,]+)\s*,\s*([^,]+)\s*,\s*([^,]+)\s*(?:,\s*//\s*(\d+))?'
    )

    # Регулярное выражение для разбиения param_name на части
    # Поддерживает русские и латинские буквы для индекса группы
    param_name_pattern = re.compile(r'([А-Яа-яA-Za-z]+)(\d+)\.(.+)')

    # Словарь для хранения параметров по группам
    parameters = {}
    current_group = None

    # Разделяем блок на строки
    lines = parameter_block.splitlines()

    for line in lines:
        # Проверяем, является ли строка началом новой группы
        group_match = group_pattern.match(line.strip())
        if group_match:
            current_group = group_match.group(2).strip()  # Название группы
            parameters[current_group] = []
            continue

        # Если текущая группа определена, ищем параметры
        if current_group:
            param_match = param_pattern.search(line)
            if param_match:
                param_name = param_match.group(1).strip()  # Имя параметра
                param_unit = param_match.group(2).strip()  # Размерность (может быть пустой)
                min_value = param_match.group(3).strip()  # Минимальное значение
                max_value = param_match.group(4).strip()  # Максимальное значение
                default_value = param_match.group(5).strip()  # Значение по умолчанию
                encoding = param_match.group(6).strip()  # Кодировка
                address = param_match.group(7).strip() if param_match.group(7) else None  # Адрес параметра

                # Разбиваем param_name на части
                param_name_match = param_name_pattern.match(param_name)
                if param_name_match:
                    group_index = param_name_match.group(1)  # Индекс группы (например, "Т")
                    param_number = param_name_match.group(2)  # Номер параметра (например, "0")
                    param_name_cleaned = param_name_match.group(3).strip()  # Имя параметра (например, "ТЕХНОЛ. РЕГ.")
                else:
                    # Если формат не совпадает, оставляем исходное имя
                    group_index = None
                    param_number = None
                    param_name_cleaned = param_name

                # Добавляем параметр в текущую группу
                parameters[current_group].append({
                    "group_index": group_index,
                    "param_number": param_number,
                    "param_name": param_name_cleaned,
                    "unit": param_unit,
                    "min_value": min_value,
                    "max_value": max_value,
                    "default_value": default_value,
                    "encoding": encoding,
                    "address": address
                })

    return parameters

def create_param_objects(file_path):
    """
    Создает массив объектов Param на основе данных из файла.

    Args:
        file_path (str): Путь к файлу.

    Returns:
        list: Список объектов Param.
    """
    # Получаем данные из парсеров
    groups = parse_groups(file_path)
    parameters = parse_parameters(file_path)

    

    # Создаем массив объектов Param
    param_objects = []

    for group in groups:
        group_number = group["group_number"]
        group_index = group["group_index"]
        group_description = group["group_description"]

        param_counter = 0  # Инициализируем счетчик 

        if group_description in parameters:
            for param in parameters[group_description]:
                param_obj = Param()

                # Заполняем поля                                
                
                param_obj.code = f"{group_number}.{param_counter}"
                param_counter += 1  # Увеличиваем счетчик

                if param["param_number"] is None:
                    param_obj.name = f"{param['param_name']}"
                else:
                    param_obj.name = f"{param['group_index']}{param['param_number']}.{param['param_name']}"
                                  
                param_obj.units = param["unit"]
                param_obj.address = str(param["address"]) if param["address"] else ""
                param_obj.view = "DEC"  
                
                # Очищаем и преобразуем числовые значения
                param_obj.min_value = clean_number(param["min_value"])
                param_obj.max_value = clean_number(param["max_value"])
                param_obj.factory_setting = clean_number(param["default_value"])
                param_obj.size = "2"

                # Определяем тип
                encoding = param["encoding"]
                if "MT_RUN" in encoding or "M_RUNS" in encoding:
                    param_obj.type = "UNION"
                elif "MT_DEC" in encoding or "MT_BIN" in encoding or "M_RMAX" in encoding:
                    if "M_SIGN" in encoding:
                        param_obj.type = "INT16"
                    else:
                        param_obj.type = "UINT16"
                elif "MT_STR" in encoding or "M_STAT" in encoding or "M_CODE" in encoding or "M_COMM" in encoding:
                    param_obj.type = "STR"
                elif "MT_DATE" in encoding or "M_DATE" in encoding:
                    param_obj.units = "ЧЧ:ММ:ГГГГ"  # Применяем формат
                    param_obj.type = "DATE"
                elif "MT_TIME" in encoding or "M_TIME" in encoding:
                    param_obj.units = "ЧЧ:ММ"  # Применяем формат
                    param_obj.type = "TIME"

                # Определяем record
                if "M_RONLY" in encoding or "M_NVM" not in encoding:
                    param_obj.record = ""
                else:
                    param_obj.record = "1"

                # Определяем coefficient
                prec_match = re.search(r'M_PREC\((\d+)\)', encoding)
                if prec_match:
                    param_obj.coefficient = str(prec_match.group(1))
                    param_obj.max_value =str(int(param_obj.max_value)/(10**int(param_obj.coefficient)))
                    param_obj.min_value =str(int(param_obj.min_value)/(10**int(param_obj.coefficient)))
                else:
                    param_obj.coefficient = ""

                # Определяем rows и description
                sadr_match = re.search(r'M_SADR\((\d+)\)', encoding)
                if sadr_match:
                    line_number = int(sadr_match.group(1))
                    strings = parse_strings(file_path, line_number)
                    param_obj.rows = "; ".join(strings)
                    param_obj.description = "; ".join(strings)

                # Определяем hidden
                if "M_HIDE" in encoding or "M_SHOW" in encoding:
                    param_obj.hidden = "1"
                else:
                    param_obj.hidden = ""

                # Добавляем объект в массив
                param_objects.append(param_obj)

    return param_objects

def parse_version(file_path):
    """
    Парсит файл и извлекает значения для ключей:
    - DEVICE_NAME
    - DEVICE_GROUP
    - VERSION
    - MODULE_VERSION
    - SUBVERSION

    Args:
        file_path (str): Путь к файлу.

    Returns:
        dict: Словарь с ключами и их значениями (все значения в виде строк).
    """
    with open(file_path, 'r', encoding='Windows-1251') as file:
        content = file.read()

    # Словарь для хранения результатов
    result = {
        "DEVICE_NAME": "",
        "DEVICE_GROUP": "",
        "VERSION": "",
        "MODULE_VERSION": "",
        "SUBVERSION": ""
    }

    # Регулярное выражение для поиска строк вида #define KEY "VALUE" или #define KEY VALUE
    pattern = re.compile(r'#define\s+({})\s+(?:"([^"]+)"|(\S+))'.format("|".join(result.keys())))

    # Ищем все совпадения
    matches = pattern.findall(content)

    # Заполняем словарь найденными значениями
    for key, value_quoted, value_unquoted in matches:
        # Если значение в кавычках, используем его, иначе используем значение без кавычек
        value = value_quoted if value_quoted else value_unquoted
        result[key] = str(value)  # Преобразуем значение в строку

    return result
        
def clean_number(value):
    """
    Очищает строку от нечисловых символов (кроме минуса и точки) и преобразует в число.
    Если значение пустое или некорректное, возвращает 0.
    """
    if not value:
        return 0
    # Удаляем все символы, кроме цифр, минуса и точки
    cleaned_value = re.sub(r"[^0-9\-.]", "", value)
    try:
        return str(cleaned_value)
    except ValueError:
        return ""
    
def folder_path(value):

    # Создаем объект Path для родительской папки
    parent_folder = Path("Description")

    # Указываем путь к вложенной папке
    subfolder = parent_folder / f"Viewer_{value['DEVICE_NAME']}_v{value['DEVICE_GROUP']}.{value['VERSION']}.{value['MODULE_VERSION']}.{value['SUBVERSION']}"

    if not os.path.exists(subfolder):
        os.makedirs(subfolder)  # Создает папку и все промежуточные папки

    return  subfolder

def main():
    """
    Главная функция программы.

    Эта функция вызывается, когда программа запускается из
    командной строки. Она содержит основную логику программы и
    отвечает за выполнение задач, указанных в аргументах
    командной строки.

    Args:
        None

    Returns:
        None
    """
    file_param_path = find_version_file()
    
    version_info = parse_version(file_param_path)
    
    print(version_info)
    
    # Search for specific files in the current directory and its subdirectories
    file_param_path = find_param_file()

    # Получаем данные
    groups = parse_groups(file_param_path)
    param_objects = create_param_objects(file_param_path)



    # Создаем Excel-файл
    
    excel_file = folder_path(version_info) / f"Viewer_{version_info['DEVICE_NAME']}_v{version_info['DEVICE_GROUP']}.{version_info['VERSION']}.{version_info['MODULE_VERSION']}.{version_info['SUBVERSION']}.xls"

    try:
        create_excel_file(excel_file, param_objects, groups)

    except Exception as e:
        print(f"Ошибка при создании файла: {e}")


    # Get the current version of the application
    version = check_app_version()
    print(version)
    
if __name__ == "__main__":
    main()

